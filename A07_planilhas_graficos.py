import matplotlib.pyplot as plt
from openpyxl import load_workbook

tabela_clientes = load_workbook(filename='tabelas/clientes_redes.xlsx')
pagamentos = tabela_clientes['aba_pagamentos']

adimplentes = 0
inadimplentes  = 0
for pagamento in pagamentos.iter_rows(min_row=1, max_row=10,min_col=1, max_col=5):
	atrasado = pagamento[3].value
	if(atrasado == 'SIM'):
		inadimplentes = inadimplentes + 1
	else:
		adimplentes = adimplentes + 1

valoresX = [1, 2]

valoresY = [adimplentes, inadimplentes]

nomes_barras = ['adimplentes', 'inadimplentes']

plt.bar(valoresX, valoresY, tick_label = nomes_barras,
		width = 0.8, color = ['green', 'red'])

plt.xlabel('situação')
plt.ylabel('quantidade')
plt.title('Gráfico inadimplência')

plt.show()
