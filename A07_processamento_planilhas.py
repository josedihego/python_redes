# xclip -selection clipboard -target image/png -out > out.png
from openpyxl import load_workbook

tabela_clientes = load_workbook(filename='tabelas/clientes_redes.xlsx')
pagamentos = tabela_clientes['aba_pagamentos']

# percorrendo os CPFs
for pagamento in pagamentos.iter_rows(min_row=1, max_row=10,
                                      min_col=1, max_col=5):
    cliente_CPF = pagamento[0].value
    print(cliente_CPF)

# percorrendo as mesalidades
for pagamento in pagamentos.iter_rows(min_row=1, max_row=10,
                                      min_col=1, max_col=5):
    cliente_CPF = pagamento[0].value
    cliente_mensalidade = pagamento[4].value
    print('O cliente de CPF: {} paga a mesalidade R$: {}'
          .format(cliente_CPF,cliente_mensalidade))

# listando clientes inadimplentes
for pagamento in pagamentos.iter_rows(min_row=1, max_row=10,
                                      min_col=1, max_col=5):
    cliente_CPF = pagamento[0].value
    atrasado = pagamento[3].value
    if(atrasado == 'SIM'):
        print('O cliente de CPF: {} esta inadimplente'
              .format(cliente_CPF))

# ordenando os clientes por data de entrada
#lista.sort(key=lambda x: x.data, reverse=True)

# media das mensalidades
soma_mens = 0
num_cli = 0
for pagamento in pagamentos.iter_rows(min_row=2, max_row=10,
                                      min_col=1, max_col=5):
    mensalidade = pagamento[4].value
    num_cli = num_cli + 1
    soma_mens = soma_mens + float(mensalidade)
print('Média de mensalidades: {:.2f} R$'.format(soma_mens/num_cli))